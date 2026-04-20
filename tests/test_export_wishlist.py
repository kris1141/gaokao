from io import BytesIO

from openpyxl import load_workbook

from app import app


def _valid_payload():
    return {
        "source": "高考分预测",
        "stream": "历史",
        "user_rank": "12345",
        "rows": [
            {
                "school_code": "10611",
                "school_name": "重庆大学",
                "major_code": "01",
                "major_name": "计算机科学与技术",
                "min_score": "612",
                "min_rank": "5120",
                "rank_ratio": "0.1133",
                "tier": "稳",
                "probability": "0.756",
            }
        ],
    }


def test_export_wishlist_success_returns_excel_file():
    client = app.test_client()

    response = client.post("/api/export-wishlist", json=_valid_payload())

    assert response.status_code == 200
    assert (
        response.headers["Content-Type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert ".xlsx" in response.headers.get("Content-Disposition", "")

    workbook = load_workbook(BytesIO(response.data))
    assert "志愿表" in workbook.sheetnames
    sheet = workbook["志愿表"]

    header_row = [cell.value for cell in sheet[1]]
    assert header_row == [
        "序号",
        "科类",
        "院校代码",
        "院校名称",
        "专业代码",
        "专业名称",
        "最低分",
        "最低位次",
        "位次比",
        "分档",
        "录取概率",
        "输入位次",
        "来源",
    ]

    assert sheet["A2"].value == 1
    assert sheet["B2"].value == "历史"
    assert sheet["D2"].value == "重庆大学"
    assert sheet["F2"].value == "计算机科学与技术"
    assert sheet["K2"].value == "75.6%"
    assert sheet["L2"].value == 12345


def test_export_wishlist_failure_when_rows_empty():
    client = app.test_client()
    payload = _valid_payload()
    payload["rows"] = []

    response = client.post("/api/export-wishlist", json=payload)

    assert response.status_code == 400
    assert response.is_json
    assert response.get_json()["ok"] is False
    assert "请至少选择 1 条志愿" in response.get_json()["error"]


def test_export_wishlist_failure_when_required_fields_missing():
    client = app.test_client()
    payload = _valid_payload()
    payload["rows"][0]["major_name"] = ""

    response = client.post("/api/export-wishlist", json=payload)

    assert response.status_code == 400
    assert response.is_json
    assert response.get_json()["ok"] is False
    assert "缺少院校或专业名称" in response.get_json()["error"]
